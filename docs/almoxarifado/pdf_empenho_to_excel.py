#!/usr/bin/env python3
"""Converte PDF Listagem de Empenhado e Nao Pago em Excel. Uso: python pdf_empenho_to_excel.py [pdf]"""

import re
import sys
from pathlib import Path

try:
    from pypdf import PdfReader
except ImportError:
    from PyPDF2 import PdfReader

import openpyxl
from openpyxl.styles import Font


def extrair_texto_pdf(caminho):
    return "\n".join((p.extract_text() or "") for p in PdfReader(caminho).pages)


def parse_valor(s):
    try:
        return float(str(s).strip().replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def extrair_contrato(hist):
    for p in [r"CONTRATO\s+N[º°]?\s*(\d+[A-Za-z]?/\d{4})", r"CONTRATO\s+(\d+[A-Za-z]?/\d{4})", r"CONTRATO\s+(\d+-\d{4})"]:
        m = re.search(p, hist, re.I)
        if m:
            return m.group(1).replace("-", "/")
    return ""


def parsear(texto):
    padrao = re.compile(r"(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(.+?)\s+(Estimativo|Global)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)")
    regs = []
    for bloco in re.split(r"(?=200[13]\.\d{2}\.\d+)", texto):
        if not bloco.strip() or "Classificacao" in bloco:
            continue
        m = padrao.search(bloco)
        if m:
            data, emp, credor, tipo, red, empenhado, liq, pag = m.groups()
            credor = credor.replace("\n", " ").strip()
            hist = re.sub(r"\s+", " ", bloco[m.end():].replace("Historico:", "").strip())[:400]
            emp_f, pag_f = parse_valor(empenhado), parse_valor(pag)
            regs.append({"numero_contrato": extrair_contrato(hist), "credor": credor, "data_empenho": data, "numero_empenho": emp, "tipo_empenho": tipo, "empenhado": emp_f, "liquidado": parse_valor(liq), "pago": pag_f, "saldo_nao_pago": emp_f - pag_f, "reduzido": parse_valor(red), "historico": hist})
    return regs


def gerar_excel(regs, saida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empenhos"
    h = ["numero_contrato", "credor", "data_empenho", "numero_empenho", "tipo_empenho", "empenhado", "liquidado", "pago", "saldo_nao_pago", "reduzido", "historico"]
    for c, x in enumerate(h, 1):
        ws.cell(1, c, x)
        ws.cell(1, c).font = Font(bold=True)
    for row, r in enumerate(regs, 2):
        for c, k in enumerate(h, 1):
            ws.cell(row, c, r.get(k, ""))
    for c in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["K"].width = 45
    wb.save(saida)
    print(f"Excel gerado: {saida} ({len(regs)} registros)")


if __name__ == "__main__":
    base = Path(__file__).parent
    pdf = base / "Listagem de Empenhado e Nao Pago.pdf" if len(sys.argv) < 2 else Path(sys.argv[1])
    if not pdf.exists():
        print(f"Arquivo nao encontrado: {pdf}")
        sys.exit(1)
    regs = parsear(extrair_texto_pdf(str(pdf)))
    if not regs:
        print("Nenhum registro extraido.")
        sys.exit(1)
    gerar_excel(regs, str(pdf.with_suffix(".xlsx")))
